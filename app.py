import streamlit as st
import pandas as pd
import numpy as np
import os
import re
import io
import warnings
import requests
import json
import csv
import ssl
from urllib3.poolmanager import PoolManager
from urllib3.util.ssl_ import create_urllib3_context
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from concurrent.futures import ThreadPoolExecutor, as_completed

# 过滤无关警告
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
warnings.filterwarnings("ignore", category=FutureWarning, module="pandas")

# 禁用不安全请求警告
requests.packages.urllib3.disable_warnings(requests.packages.urllib3.exceptions.InsecureRequestWarning)

# 创建自定义SSL上下文
class CustomSSLAdapter(requests.adapters.HTTPAdapter):
    def init_poolmanager(self, connections, maxsize, block=False):
        context = create_urllib3_context()
        # 允许较旧的SSL/TLS版本
        context.min_version = ssl.TLSVersion.TLSv1_2
        context.check_hostname = False
        context.verify_mode = ssl.CERT_NONE
        
        self.poolmanager = PoolManager(
            num_pools=connections,
            maxsize=maxsize,
            block=block,
            ssl_context=context
        )

# -------------------------- 全局常量 & 样式定义 --------------------------
# 科目代码映射，将从API获取
SUBJECT_CODE_MAP = {
    
}

header_style = NamedStyle(name="header_style")
header_style.font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
header_style.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_style.alignment = Alignment(horizontal='center', vertical='center')
header_style.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

data_style = NamedStyle(name="data_style")
data_style.font = Font(name='微软雅黑', size=10)
data_style.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
data_style.alignment = Alignment(horizontal='left', vertical='center')

center_data_style = NamedStyle(name="center_data_style")
center_data_style.font = Font(name='微软雅黑', size=10)
center_data_style.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
center_data_style.alignment = Alignment(horizontal='center', vertical='center')

# -------------------------- 页面基础设置 & 状态初始化 --------------------------
st.set_page_config(page_title="多科目异常追踪系统", page_icon="🎯", layout="wide")

if 'analysis_completed' not in st.session_state:
    st.session_state.analysis_completed = False
    st.session_state.excel_bytes = None
    st.session_state.report_sheets = {}
    st.session_state.diff_df = pd.DataFrame()
    st.session_state.img_mapping = {}
    st.session_state.enable_viewer = False
    st.session_state.current_idx = 0
    st.session_state.current_filter = "全部显示"
    st.session_state.data_changed = False
    st.session_state.token = None
    st.session_state.schools = []
    st.session_state.selected_school = None
    st.session_state.exams = []
    st.session_state.selected_exam = None
    st.session_state.courses = []
    st.session_state.selected_courses = []
    st.session_state.all_merged_data = pd.DataFrame()

# -------------------------- API 相关函数 --------------------------
def login(username, password):
    """
    登录系统获取token
    :param username: 登录手机号
    :param password: 登录密码
    :return: token字符串
    """
    login_url = "https://www.txdat.net/auth"
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "Referer": "https://www.txdat.net/login",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0 Safari/537.36"
    }
    
    login_data = {
        "Username": username,
        "Password": password
    }
    
    response = requests.post(
        url=login_url,
        headers=headers,
        data=login_data,
        allow_redirects=False,
        verify=False
    )
    
    # 提取Token
    token = None
    set_cookie_header = response.headers.get("set-cookie", "")
    
    # 处理set_cookie_header可能是列表的情况
    if isinstance(set_cookie_header, list):
        set_cookie_header = ", ".join(set_cookie_header)
    
    for cookie in set_cookie_header.split(", "):
        if cookie.startswith("token="):
            # 修复分割逻辑
            token_part = cookie.split(";").pop(0)
            token = token_part.split("=")[1]
            break
    
    return token

def call_api(token, endpoint, method="GET", data=None, is_json=False):
    """
    调用API接口
    :param token: 认证token
    :param endpoint: API端点
    :param method: 请求方法
    :param data: 请求数据
    :param is_json: 是否以JSON格式发送数据
    :return: 响应数据
    """
    api_url = f"https://api.txdat.net{endpoint}"
    headers = {
        "Authorization": f"Bearer {token}",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0 Safari/537.36"
    }
    
    if is_json:
        headers["Content-Type"] = "application/json"
    
    # 创建会话并使用自定义SSL适配器
    session = requests.Session()
    session.mount('https://', CustomSSLAdapter())
    
    try:
        if method.upper() == "GET":
            response = session.get(api_url, headers=headers, params=data, verify=False, timeout=30)
        elif method.upper() == "POST":
            if is_json:
                response = session.post(api_url, headers=headers, json=data, verify=False, timeout=30)
            else:
                response = session.post(api_url, headers=headers, data=data, verify=False, timeout=30)
        
        response.raise_for_status()  # 检查响应状态码
        return response.json()
    except requests.exceptions.RequestException as e:
        st.error(f"API调用失败: {str(e)}")
        return None

def get_schools(token):
    """
    获取学校列表
    :param token: 认证token
    :return: 学校列表
    """
    return call_api(token, "/school/schools")

def get_exams(token, schoolid, status=-1):
    """
    获取考试列表
    :param token: 认证token
    :param schoolid: 学校ID
    :param status: 状态 (-1: 全部, 0: 未结束, 1: 已结束)
    :return: 考试列表
    """
    return call_api(token, "/exam/exams", data={"schoolid": schoolid, "status": status})

def get_exam_courses(token, examgroup):
    """
    获取考试科目列表
    :param token: 认证token
    :param examgroup: 考试组号
    :return: 科目列表
    """
    return call_api(token, "/exam/courses", data={"examgroup": examgroup})

def get_system_courses(token, schoolid):
    """
    获取系统内置的所有科目信息
    :param token: 认证token
    :param schoolid: 学校ID
    :return: 科目列表
    """
    return call_api(token, "/course/list", data={"schoolid": schoolid})

def get_exam_students(token, examid, limit=10000):
    """
    获取考试学生列表
    :param token: 认证token
    :param examid: 考试ID
    :param limit: 每页数量
    :return: 学生列表
    """
    # 分页获取所有学生数据
    all_students = []
    offset = 0
    
    while True:
        # 使用浏览器中的参数
        students = call_api(token, "/student/exam/query", data={
            "sort": "code",
            "order": "asc",
            "offset": offset,
            "limit": limit,
            "examid": examid,
            "groupid": 0,
            "column": "code",
            "value": "",
            "isscan": -1,
            "isupload": -1
        })
        
        if students is None:
            break
        
        # 处理students可能是字典的情况
        if isinstance(students, dict):
            # 尝试从字典中提取学生列表
            if 'students' in students:
                students = students['students']
            elif 'data' in students:
                students = students['data']
            elif 'items' in students:
                students = students['items']
            elif 'list' in students:
                students = students['list']
            else:
                # 尝试查找可能包含学生数据的键
                for key, value in students.items():
                    if isinstance(value, list):
                        students = value
                        break
                else:
                    break
        
        if not isinstance(students, list):
            break
        
        if not students:
            break
        
        all_students.extend(students)
        offset += limit
    
    return all_students

def get_image_links(token, examid):
    """
    获取考试相关图片链接
    :param token: 认证token
    :param examid: 考试ID
    :return: 图片链接列表，每个元素包含link和upload_path
    """
    try:
        # 下载包含链接的txt文件
        url = f"https://api.txdat.net/file/text/downimg"
        params = {
            "token": token,
            "examid": examid
        }
        
        # 创建会话并使用自定义SSL适配器
        session = requests.Session()
        session.mount('https://', CustomSSLAdapter())
        
        response = session.get(url, params=params, verify=False, timeout=30)
        response.raise_for_status()
        
        # 读取txt文件内容，获取下载链接
        content = response.text.strip()
        
        # 处理下载链接，清理特殊字符和额外信息
        links_info = []
        for line in content.split('\n'):
            # 清理链接
            line = line.strip()
            if not line:
                continue
            
            # 分割链接和上传路径
            parts = line.split('\t')
            if len(parts) >= 2:
                link = parts[0].strip()
                upload_path = parts[1].strip()
                # 清理URL中的特殊字符
                link = link.replace('%09', '').replace('%0D', '')
                links_info.append({"link": link, "upload_path": upload_path})
            elif line.startswith('http://') or line.startswith('https://'):
                # 只保留URL部分，去除回车符
                link = line.split('\r')[0]
                # 清理URL中的特殊字符
                link = link.replace('%09', '').replace('%0D', '')
                links_info.append({"link": link, "upload_path": ""})
        
        if not links_info:
            st.warning("未找到有效的下载链接")
            return []
        st.success(f"获取到 {len(links_info)} 个图片链接")
        # 静默处理，不输出链接信息
        return links_info
    except Exception as e:
        st.error(f"获取图片链接失败: {str(e)}")
        return []

# -------------------------- 核心工具函数 --------------------------
def process_student_data(students, course_name):
    """
    处理学生数据，转换为DataFrame
    :param students: 学生列表
    :param course_name: 科目名称
    :return: 处理后的DataFrame
    """
    if not students:
        return pd.DataFrame()
    
    # 转换为DataFrame
    df = pd.DataFrame(students)
    
    # 重命名和处理字段
    df = df.rename(columns={
        'code': '考号',
        'name': '姓名',
        'schoolname': '学校',
        'classroom': '班级',
        'studid': '学号',
        'scandone': '扫描状态'
    })
    
    # 处理扫描状态
    scan_status_map = {'True': '已扫', 'False': '未扫', '1': '已扫', '0': '未扫', '是': '已扫', '否': '未扫', '已扫': '已扫', '未扫': '未扫', '': '未扫', '未记录': '未记录'}
    df['扫描状态'] = df['扫描状态'].astype(str).str.strip().replace(scan_status_map).fillna('未记录')
    
    # 添加科目列
    df['科目'] = course_name
    
    # 确保必要的列存在
    required_cols = ['考号', '姓名', '学校', '班级', '学号', '扫描状态', '科目']
    for col in required_cols:
        if col not in df.columns:
            df[col] = '未记录'
    
    return df[required_cols]

def generate_student_list_data(all_merged_data):
    student_list = all_merged_data.groupby('考号').agg(
        姓名=('姓名', 'first'), 学校=('学校', 'first'), 班级=('班级', 'first'),
        科目=('科目', lambda x: '; '.join(x.unique()) if not x.empty else '无'),
        涉及科目数=('科目', 'nunique')
    ).reset_index()
    return student_list[['考号', '姓名', '学校', '班级', '科目', '涉及科目数']]

def classify_scan_status(all_merged_data):
    status_summary = all_merged_data.groupby('考号').agg(
        学校=('学校', 'first'), 姓名=('姓名', 'first'), 学号=('学号', 'first'),
        班级=('班级', 'first'), 科目=('科目', lambda x: ', '.join(x.unique()) if not x.empty else '无')
    ).reset_index()
    
    is_scanned = all_merged_data['扫描状态'] == '已扫'
    is_unscanned = all_merged_data['扫描状态'] == '未扫'
    is_unrecorded = all_merged_data['扫描状态'] == '未记录'

    scanned_subj = all_merged_data[is_scanned].groupby('考号')['科目'].apply(lambda x: '; '.join(x)).rename('已扫科目')
    unscanned_subj = all_merged_data[is_unscanned].groupby('考号')['科目'].apply(lambda x: '; '.join(x)).rename('未扫_纯')
    unrecorded_subj = all_merged_data[is_unrecorded].groupby('考号')['科目'].apply(lambda x: '; '.join([f"{s}(未记录)" for s in x])).rename('未扫_记录')

    df_subj = pd.DataFrame(index=status_summary['考号']).join(scanned_subj).join(unscanned_subj).join(unrecorded_subj)
    df_subj['已扫科目'] = df_subj['已扫科目'].fillna('无')
    s1, s2 = df_subj['未扫_纯'].fillna(''), df_subj['未扫_记录'].fillna('')
    sep = np.where((s1 != '') & (s2 != ''), '; ', '')
    df_subj['未扫科目'] = (s1 + sep + s2).replace('', '无')

    status_summary = status_summary.merge(df_subj[['已扫科目', '未扫科目']], on='考号', how='left')

    counts = all_merged_data.pivot_table(index='考号', columns='扫描状态', values='科目', aggfunc='count', fill_value=0, observed=False)
    for col in ['已扫', '未扫', '未记录']:
        if col not in counts.columns:
            counts[col] = 0
            
    total_valid = counts['已扫'] + counts['未扫']
    conditions = [total_valid == 0, (counts['已扫'] > 0) & (counts['未扫'] == 0), (counts['未扫'] > 0) & (counts['已扫'] == 0)]
    counts['扫描状态分类'] = np.select(conditions, ['状态未记录', '全已扫', '全未扫'], default='状态差异')
    status_summary = status_summary.merge(counts[['扫描状态分类']], on='考号', how='left')
    status_summary = status_summary.merge(all_merged_data.groupby('考号')['扫描状态'].apply(lambda x: list(x.unique())).rename('扫描状态'), on='考号', how='left')
    status_summary['涉及科目数'] = status_summary['科目'].str.count(',') + 1
    
    final_cols = ['考号', '学校', '姓名', '学号', '班级', '科目', '涉及科目数', '已扫科目', '未扫科目', '扫描状态', '扫描状态分类']
    return {
        '全已扫': status_summary[status_summary['扫描状态分类'] == '全已扫'][final_cols],
        '全未扫': status_summary[status_summary['扫描状态分类'] == '全未扫'][final_cols],
        '状态差异': status_summary[status_summary['扫描状态分类'] == '状态差异'][final_cols],
        '状态未记录': status_summary[status_summary['扫描状态分类'] == '状态未记录'][final_cols]
    }

def create_scan_pivot_table(all_merged_data):
    pivot_data = all_merged_data.copy()
    pivot_data['是否已扫'] = pivot_data['扫描状态'] == '已扫'
    
    total_pivot = pd.pivot_table(pivot_data, index='学校', columns='科目', values='考号', aggfunc='count', fill_value=0, observed=False)
    scanned_pivot = pd.pivot_table(pivot_data[pivot_data['是否已扫']], index='学校', columns='科目', values='考号', aggfunc='count', fill_value=0, observed=False)
    
    all_schools, all_subjects = total_pivot.index.union(scanned_pivot.index), total_pivot.columns.union(scanned_pivot.columns)
    total_pivot, scanned_pivot = total_pivot.reindex(index=all_schools, columns=all_subjects, fill_value=0), scanned_pivot.reindex(index=all_schools, columns=all_subjects, fill_value=0)
    
    result_pivot = scanned_pivot.astype(str) + '/' + total_pivot.astype(str)
    percentage_pivot = pd.DataFrame(np.where(total_pivot > 0, (scanned_pivot / total_pivot) * 100, 0.0), index=all_schools, columns=all_subjects).applymap(lambda x: f"{x:.1f}%")
    
    row_scanned, row_total = scanned_pivot.sum(axis=1), total_pivot.sum(axis=1)
    result_pivot['总计'] = row_scanned.astype(str) + '/' + row_total.astype(str)
    percentage_pivot['总计'] = [f"{x:.1f}%" for x in np.where(row_total > 0, (row_scanned / row_total) * 100, 0.0)]
    
    col_scanned, col_total = scanned_pivot.sum(axis=0), total_pivot.sum(axis=0)
    col_scanned['总计'], col_total['总计'] = col_scanned.sum(), col_total.sum()
    
    result_pivot.loc['总计'] = col_scanned.astype(str) + '/' + col_total.astype(str)
    percentage_pivot.loc['总计'] = [f"{x:.1f}%" for x in np.where(col_total > 0, (col_scanned / col_total) * 100, 0.0)]
    
    result_pivot = result_pivot.rename_axis('学校').reset_index()
    percentage_pivot = percentage_pivot.rename_axis('学校').reset_index()
    
    return {'数量透视表': result_pivot, '百分比透视表': percentage_pivot}

def parse_image_mappings(links_info_list, course_names):
    """
    解析图片映射信息
    :param links_info_list: 图片链接列表
    :param course_names: 科目名称列表
    :return: 图片映射字典
    """
    mapping = {}
    for links_info, course_name in zip(links_info_list, course_names):
        for item in links_info:
            link = item.get('link', '')
            upload_path = item.get('upload_path', '')
            if link and upload_path:
                # 从上传路径中提取学生信息
                parts = upload_path.split('\\')
                if len(parts) >= 3 and '(1)' in parts[-1]:
                    student_dir = parts[-2]
                    student_id = student_dir.split('(')[0]
                    mapping[(student_id, course_name)] = link
    return mapping

def set_excel_cell_style_optimized(ws):
    for cell in ws[1]:
        cell.style = header_style
    max_row, max_col, headers = ws.max_row, ws.max_column, [cell.value for cell in ws[1]]
    center_cols_idx = {i for i, h in enumerate(headers) if h in ['涉及科目数', '扫描状态分类', '考号', '学号', '处理进度']}
    if '学校' in headers and max_row >= 2:
        center_cols_idx = set(range(max_col))
    
    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
        for idx, cell in enumerate(row):
            if headers[idx] == '处理进度' and cell.value == '已标记':
                cell.font = Font(name='微软雅黑', size=10, color='008000', bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            else:
                cell.style = center_data_style if idx in center_cols_idx else data_style
    
    for col in ws.columns:
        col_letter, header_val = col[0].column_letter, str(col[0].value) if col[0].value else ""
        max_len = max([len(str(cell.value)) for cell in col[:100] if cell.value] or [0])
        width = min(max_len + 4, 40) if '科目' in header_val and '数' not in header_val else min(max_len + 3, 20) if '总计' in header_val or '学校' in header_val else min(max_len + 3, 15)
        ws.column_dimensions[col_letter].width = width

def generate_latest_excel():
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # 只包含需要的四张表
        # 1. 名单数据
        if '名单数据' in st.session_state.report_sheets:
            df = st.session_state.report_sheets['名单数据']
            df.to_excel(writer, sheet_name='名单', index=False)
            set_excel_cell_style_optimized(writer.sheets['名单'])
        
        # 2. 全已扫
        if '全已扫' in st.session_state.report_sheets:
            df = st.session_state.report_sheets['全已扫']
            df.to_excel(writer, sheet_name='全已扫', index=False)
            set_excel_cell_style_optimized(writer.sheets['全已扫'])
        
        # 3. 全未扫
        if '全未扫' in st.session_state.report_sheets:
            df = st.session_state.report_sheets['全未扫']
            df.to_excel(writer, sheet_name='全未扫', index=False)
            set_excel_cell_style_optimized(writer.sheets['全未扫'])
        
        # 4. 差异表
        if not st.session_state.diff_df.empty:
            df = st.session_state.diff_df
            df.to_excel(writer, sheet_name='状态差异', index=False)
            set_excel_cell_style_optimized(writer.sheets['状态差异'])
    st.session_state.excel_bytes = output.getvalue()

# -------------------------- Streamlit 网页前端逻辑 --------------------------
st.set_page_config(page_title="多科目异常追踪系统", page_icon="🎯", layout="wide")
st.title("🎯 多科目异常追踪与极速打标系统")
st.markdown("不仅永久记忆您的标记，更引入了**「快捷键翻页」「快捷标签秒杀」与「按校批量处理」**三大效率神器！")

# 注入键盘事件监听（前端黑科技）
st.markdown("""
<script>
    if (!window.parent.customNavListener) {
        window.parent.document.addEventListener('keydown', function(e) {
            if (e.key === 'ArrowLeft') {
                const btn = Array.from(window.parent.document.querySelectorAll('button')).find(el => el.innerText.includes('上一个'));
                if (btn) btn.click();
            } else if (e.key === 'ArrowRight') {
                const btn = Array.from(window.parent.document.querySelectorAll('button')).find(el => el.innerText.includes('下一个'));
                if (btn) btn.click();
            }
        });
        window.parent.customNavListener = true;
    }
</script>
""", unsafe_allow_html=True)

# 登录部分
if not st.session_state.token:
    st.subheader("🔐 登录系统")
    login_choice = st.radio("选择登录方式:", ["使用用户名密码登录", "直接输入token"], horizontal=True)
    
    if login_choice == "使用用户名密码登录":
        username = st.text_input("登录手机号:")
        password = st.text_input("登录密码:", type="password")
        
        if st.button("登录", type="primary"):
            if not username or not password:
                st.warning("请输入用户名和密码")
            else:
                with st.spinner("正在登录..."):
                    token = login(username, password)
                    if token:
                        st.session_state.token = token
                        st.success("登录成功，获取到token")
                        st.rerun()
                    else:
                        st.error("登录失败，无法获取token")
    else:
        token = st.text_input("请输入token:")
        
        if st.button("确认", type="primary"):
            if not token:
                st.warning("token不能为空")
            else:
                st.session_state.token = token
                st.success("已输入token")
                st.rerun()
else:
    # 上传历史标记报告
    st.subheader("📤 上传历史标记报告 (可选)")
    history_file = st.file_uploader("上传上次导出的本系统标记报告", type=['xlsx', 'xls'], accept_multiple_files=False)
    
    # 获取学校列表
    if not st.session_state.schools:
        with st.spinner("正在获取学校列表..."):
            schools = get_schools(st.session_state.token)
            if schools:
                st.session_state.schools = schools
            else:
                st.error("未获取到学校列表")
                st.session_state.token = None
                st.rerun()
    
    # 选择学校
    if st.session_state.schools:
        school_names = [school['schoolname'] for school in st.session_state.schools]
        selected_school_name = st.selectbox("选择学校:", school_names)
        selected_school = next((school for school in st.session_state.schools if school['schoolname'] == selected_school_name), None)
        
        if selected_school:
            # 检查是否切换了学校
            if st.session_state.selected_school is None or st.session_state.selected_school['schoolid'] != selected_school['schoolid']:
                # 重置考试相关状态
                st.session_state.selected_school = selected_school
                st.session_state.exams = []
                st.session_state.selected_exam = None
                st.session_state.courses = []
                st.session_state.selected_courses = []
                st.session_state.analysis_completed = False
                
                # 获取系统内置科目信息，更新SUBJECT_CODE_MAP
                with st.spinner("正在获取系统科目信息..."):
                    system_courses = get_system_courses(st.session_state.token, selected_school['schoolid'])
                    if system_courses:
                        # 构建科目代码映射
                        new_subject_map = {}
                        for course in system_courses:
                            # 将courseid转换为两位数字符串
                            course_id_str = str(course['courseid']).zfill(2)
                            new_subject_map[course_id_str] = course['coursename']
                        # 更新全局变量
                        SUBJECT_CODE_MAP = new_subject_map
                        st.success("成功获取系统科目信息")
                    else:
                        st.warning("未获取到系统科目信息，使用默认科目映射")
            
            # 获取考试列表
            if not st.session_state.exams:
                with st.spinner("正在获取考试列表..."):
                    exams = get_exams(st.session_state.token, selected_school['schoolid'])
                    if exams:
                        st.session_state.exams = exams
                    else:
                        st.error("未获取到考试列表")
            
            # 选择考试
            if st.session_state.exams:
                exam_names = [exam['examname'] for exam in st.session_state.exams]
                selected_exam_name = st.selectbox("选择考试:", exam_names)
                selected_exam = next((exam for exam in st.session_state.exams if exam['examname'] == selected_exam_name), None)
                
                if selected_exam:
                    # 检查是否切换了考试
                    if st.session_state.selected_exam is None or st.session_state.selected_exam['examgroup'] != selected_exam['examgroup']:
                        # 重置科目相关状态
                        st.session_state.selected_exam = selected_exam
                        st.session_state.courses = []
                        st.session_state.selected_courses = []
                        st.session_state.analysis_completed = False
                    
                    # 获取考试科目
                    if not st.session_state.courses:
                        with st.spinner("正在获取考试科目..."):
                            courses = get_exam_courses(st.session_state.token, selected_exam['examgroup'])
                            if courses:
                                st.session_state.courses = courses
                            else:
                                st.error("未获取到考试科目")
                    
                    # 选择科目
                    if st.session_state.courses:
                        course_names = [course['coursename'] for course in st.session_state.courses]
                        selected_course_indices = st.multiselect("选择科目:", range(len(course_names)), format_func=lambda x: course_names[x])
                        
                        if selected_course_indices:
                            selected_courses = [st.session_state.courses[idx] for idx in selected_course_indices]
                            st.session_state.selected_courses = selected_courses
                            
                            # 获取数据按钮
                            if st.button("🚀 开始获取数据并分析", type="primary", use_container_width=True):
                                with st.spinner("正在获取数据并分析..."):
                                    # 批量获取学生数据
                                    all_data = []
                                    links_info_list = []
                                    course_names_selected = []
                                    
                                    for course in selected_courses:
                                        st.write(f"正在处理科目: {course['coursename']}")
                                        # 获取学生数据
                                        students = get_exam_students(st.session_state.token, course['examid'])
                                        if students:
                                            # 处理学生数据
                                            df = process_student_data(students, course['coursename'])
                                            all_data.append(df)
                                            # 获取图片链接
                                            links_info = get_image_links(st.session_state.token, course['examid'])
                                            links_info_list.append(links_info)
                                            course_names_selected.append(course['coursename'])
                                        else:
                                            st.warning(f"未获取到 {course['coursename']} 的学生数据")
                                    
                                    # 合并所有数据
                                    if all_data:
                                        all_merged = pd.concat(all_data, ignore_index=True)
                                        st.session_state.all_merged_data = all_merged
                                        
                                        # 分类扫描状态
                                        classification_result = classify_scan_status(all_merged)
                                        
                                        # 生成学生名单数据
                                        st.session_state.report_sheets['名单数据'] = generate_student_list_data(all_merged)
                                        
                                        # 保存分类结果到report_sheets
                                        for sheet_name, df in classification_result.items():
                                            if not df.empty:
                                                st.session_state.report_sheets[sheet_name] = df
                                        
                                        # 处理状态差异数据
                                        new_diff_df = classification_result['状态差异'].copy()
                                        if not new_diff_df.empty:
                                            new_diff_df.insert(0, '处理进度', '未处理')
                                            new_diff_df.insert(1, '处理备注', '')
                                        
                                        # 继承历史标记
                                        inherited_count = 0
                                        if history_file and not new_diff_df.empty:
                                            try:
                                                hist_df = pd.read_excel(history_file, sheet_name='状态差异')
                                                if '考号' in hist_df.columns and '处理进度' in hist_df.columns:
                                                    hist_status = hist_df.set_index(hist_df['考号'].astype(str))['处理进度'].to_dict()
                                                    hist_remark = hist_df.set_index(hist_df['考号'].astype(str))['处理备注'].to_dict() if '处理备注' in hist_df.columns else {}
                                                    new_diff_df['处理进度'] = new_diff_df.apply(lambda r: hist_status.get(str(r['考号']), '未处理'), axis=1)
                                                    new_diff_df['处理备注'] = new_diff_df.apply(lambda r: hist_remark.get(str(r['考号']), '') if pd.notna(hist_remark.get(str(r['考号']), '')) else '', axis=1)
                                                    inherited_count = len(new_diff_df[new_diff_df['处理进度'] == '已标记'])
                                            except Exception as e:
                                                st.warning(f"⚠️ 提取历史继承数据时出错: {e}")
                                        
                                        st.session_state.diff_df = new_diff_df
                                        
                                        # 显示继承结果
                                        if inherited_count > 0:
                                            st.success(f"🎉 成功继承了 {inherited_count} 名考生的历史标记！")
                                        
                                        # 解析图片映射
                                        if links_info_list:
                                            st.session_state.img_mapping = parse_image_mappings(links_info_list, course_names_selected)
                                            st.session_state.enable_viewer = True
                                        else:
                                            st.session_state.enable_viewer = False
                                            st.session_state.img_mapping = {}
                                        
                                        st.session_state.analysis_completed = True
                                        st.session_state.current_idx = 0
                                        st.session_state.current_filter = "全部显示"
                                        st.session_state.data_changed = False
                                        generate_latest_excel()
                                        st.success("数据获取和分析完成！")
                                    else:
                                        st.error("未获取到任何学生数据")

# -------------------------- 沉浸式双核处理台 --------------------------
if st.session_state.analysis_completed:
    st.divider()
    diff_df = st.session_state.diff_df
    
    if diff_df.empty:
        st.success("🎉 太棒了！本次最新数据中没有任何状态差异的考生。")
    else:
        total_diff = len(diff_df)
        processed_count = len(diff_df[diff_df['处理进度'] == '已标记'])
        pending_count = total_diff - processed_count
        
        st.header("💻 异常名单智能处理台")
        col1, col2, col3 = st.columns(3)
        col1.metric("⚠️ 差异大名单总人数", f"{total_diff} 人")
        col2.metric("✅ 已知情 / 已标记", f"{processed_count} 人")
        col3.metric("⏳ 尚未查明待处理", f"{pending_count} 人")
        
        # -------------------------- 极速下载控制台 --------------------------
        st.markdown("### 💾 数据保存与导出")
        dl_col1, dl_col2 = st.columns(2)
        with dl_col1:
            if st.session_state.get('data_changed', False):
                if st.button("🔄 1. 进度已更新，点击打包最新文件", use_container_width=True):
                    with st.spinner("📦 正在极速生成最新 Excel..."):
                        generate_latest_excel()
                        st.session_state.data_changed = False
                        st.rerun()
            else:
                st.button("✅ 1. 生成功能已是最新状态", disabled=True, use_container_width=True)
                
        with dl_col2:
            if not st.session_state.get('data_changed', False) and st.session_state.excel_bytes:
                st.download_button(
                    label="📥 2. 一键下载最新 Excel 报告 (作为下次继承凭证)",
                    data=st.session_state.excel_bytes,
                    file_name="多科目状态追踪与处理报告.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True
                )
            else:
                st.button("📥 2. 请先点击左侧按钮打包数据", disabled=True, use_container_width=True)

        st.markdown("---")
        
        # -------------------------- 双模式工作台 Tab --------------------------
        tab1, tab2 = st.tabs(["🎯 逐个精细核对 (支持快捷键)", "🗂️ 按学校批量打标"])

        # ========================== TAB 1: 逐个核对模式 ==========================
        with tab1:
            selected_filter = st.radio(
                "🔎 **快速筛选视图：**", ["全部显示", "🔴 只看未处理", "🟢 只看已标记"], horizontal=True, 
                index=["全部显示", "🔴 只看未处理", "🟢 只看已标记"].index(st.session_state.current_filter)
            )

            if selected_filter != st.session_state.current_filter:
                st.session_state.current_filter = selected_filter
                st.session_state.current_idx = 0
                st.rerun()

            if selected_filter == "🔴 只看未处理":
                display_df = diff_df[diff_df['处理进度'] == '未处理']
            elif selected_filter == "🟢 只看已标记":
                display_df = diff_df[diff_df['处理进度'] == '已标记']
            else:
                display_df = diff_df

            if display_df.empty:
                st.info(f"✨ 当前视图下 ({selected_filter}) 暂无需要处理的考生。")
            else:
                options = []
                for _, row in display_df.iterrows():
                    status_icon = "🟢 [已标记]" if str(row.get('处理进度')) == '已标记' else "🔴 [未处理]"
                    options.append(f"{status_icon} {row['考号']} | {row['姓名']} | {row['学校']} {row['班级']}")

                if st.session_state.current_idx >= len(options):
                    st.session_state.current_idx = max(0, len(options) - 1)

                st.write("🔍 **快速定位或快捷键切换 (支持 ← 左方向键 / 右方向键 →)：**")
                nav_col1, nav_col2, nav_col3 = st.columns([1, 8, 1])
                with nav_col1:
                    if st.button("⬅️ 上一个", disabled=(st.session_state.current_idx == 0), use_container_width=True):
                        st.session_state.current_idx -= 1
                        st.rerun()
                with nav_col2:
                    selected_option = st.selectbox("隐藏", options, index=st.session_state.current_idx, label_visibility="collapsed")
                    actual_idx = options.index(selected_option)
                    if actual_idx != st.session_state.current_idx:
                        st.session_state.current_idx = actual_idx
                        st.rerun()
                with nav_col3:
                    if st.button("下一个 ➡️", disabled=(st.session_state.current_idx == len(options) - 1), use_container_width=True):
                        st.session_state.current_idx += 1
                        st.rerun()
                
                st.markdown("<br>", unsafe_allow_html=True)
                
                student_id = selected_option.split(" | ")[0].split("] ")[1].strip()
                student_name = selected_option.split(" | ")[1].strip()
                matched_rows = display_df[display_df['考号'].astype(str) == student_id]
                
                if not matched_rows.empty:
                    row_data = matched_rows.iloc[0]
                    current_remark = str(row_data.get('处理备注', ''))
                    scanned_subjs = str(row_data['已扫科目']).split('; ')
                    unscanned_subjs = str(row_data['未扫科目'])
                    
                    view_col, action_col = st.columns([6, 4])
                    
                    with view_col:
                        st.markdown(f"**当前查验：** `{student_name} ({student_id})` 　|　 ❌ **差异科目：** `{unscanned_subjs}`")
                        if not st.session_state.enable_viewer:
                            st.info("ℹ️ 未获取到图片链接，如已知晓原因，请直接在右侧填写备注并标记。")
                        else:
                            valid_images = []
                            for subj in scanned_subjs:
                                url = st.session_state.img_mapping.get((student_id, subj))
                                if url:
                                    valid_images.append((subj, url))
                            
                            if valid_images:
                                img_cols = st.columns(len(valid_images) if len(valid_images) <= 2 else 2)
                                for idx, (subj, url) in enumerate(valid_images):
                                    with img_cols[idx % 2]:
                                        st.markdown(f"📄 **{subj}**")
                                        st.markdown(f'''<a href="{url}" target="_blank"><img src="{url}" style="width:100%; border-radius:4px; border:1px solid #ccc;"/></a>''', unsafe_allow_html=True)
                            else:
                                st.warning("⚠️ 暂无匹配的试卷第一页图片。")
                    
                    with action_col:
                        with st.container(border=True):
                            st.subheader("⚡ 快捷标签一键秒杀")
                            st.caption("点击下方按钮，将自动填入备注并跳到下一个人")
                            
                            def save_and_go_next(remark):
                                idx_to_update = st.session_state.diff_df.index[st.session_state.diff_df['考号'].astype(str) == student_id].tolist()
                                if idx_to_update:
                                    st.session_state.diff_df.at[idx_to_update[0], '处理进度'] = '已标记'
                                    st.session_state.diff_df.at[idx_to_update[0], '处理备注'] = remark
                                    st.session_state.data_changed = True
                                    if selected_filter == "🔴 只看未处理":
                                        pass
                                    else:
                                        if st.session_state.current_idx < len(options) - 1:
                                            st.session_state.current_idx += 1
                                    st.rerun()

                            tag_col1, tag_col2 = st.columns(2)
                            if tag_col1.button("🚨 缺考", use_container_width=True):
                                save_and_go_next("缺考")
                            if tag_col2.button("📝 白卷", use_container_width=True):
                                save_and_go_next("白卷")
                            if tag_col1.button("✂️ 条码破损", use_container_width=True):
                                save_and_go_next("条码破损")
                            if tag_col2.button("🚫 走错考场", use_container_width=True):
                                save_and_go_next("走错考场")
                            if st.button("⏪ 撤销标记 (退回未处理)", use_container_width=True):
                                idx_to_update = st.session_state.diff_df.index[st.session_state.diff_df['考号'].astype(str) == student_id].tolist()
                                if idx_to_update:
                                    st.session_state.diff_df.at[idx_to_update[0], '处理进度'] = '未处理'
                                    st.session_state.diff_df.at[idx_to_update[0], '处理备注'] = ''
                                    st.session_state.data_changed = True
                                    if selected_filter == "🟢 只看已标记":
                                        pass
                                    else:
                                        if st.session_state.current_idx < len(options) - 1:
                                            st.session_state.current_idx += 1
                                    st.rerun()

                            st.divider()
                            st.subheader("✏️ 或者手动输入其他原因")
                            custom_remark = st.text_input("差异原因：", value=current_remark if current_remark != 'nan' else '')
                            if st.button("✅ 手动保存并下一个", type="primary", use_container_width=True):
                                save_and_go_next(custom_remark)

        # ========================== TAB 2: 按学校批量处理 ==========================
        with tab2:
            st.subheader("🏢 按学校快速批量标记")
            st.caption("对于因为未收齐等原因导致的某个学校群体性异常，您可以在这里一键全选标记。")
            
            # 只提取未处理的数据进行批量操作
            unprocessed_df = diff_df[diff_df['处理进度'] == '未处理']
            
            if unprocessed_df.empty:
                st.success("🎉 太棒了，当前没有任何「未处理」的学生需要批量打标！")
            else:
                # 计算每个学校的未处理学生数量
                school_counts = unprocessed_df['学校'].value_counts()
                # 创建显示格式：学校 | 人数
                school_options = [f"{school} | {school_counts[school]}人" for school in school_counts.index]
                selected_option = st.selectbox("1️⃣ 请选择要批量处理的学校：", school_options)
                # 提取学校名称
                selected_school = selected_option.split(' | ')[0]
                
                school_df = unprocessed_df[unprocessed_df['学校'] == selected_school]
                school_count = len(school_df)
                
                st.info(f"📊 数据检测：该学校目前共有 **{school_count}** 名未处理的异常考生。")
                
                batch_remark = st.text_input("2️⃣ 请输入批量统一的异常原因：", placeholder="例如：该考场试卷未回收...")
                
                if st.button(f"⚡ 确认将这 {school_count} 人全部标记为上述原因", type="primary"):
                    if not batch_remark:
                        st.warning("⚠️ 请先输入异常原因！")
                    else:
                        # 获取这批学生在总表中的真实索引并统一修改
                        batch_indexes = school_df.index
                        st.session_state.diff_df.loc[batch_indexes, '处理进度'] = '已标记'
                        st.session_state.diff_df.loc[batch_indexes, '处理备注'] = batch_remark
                        st.session_state.data_changed = True
                        st.session_state.current_idx = 0  # 重置单兵模式的指针
                        
                        st.toast(f"✅ 成功将 {selected_school} 的 {school_count} 名考生批量标记！")
                        st.rerun()
